"""Export phiếu supplier ra template.xlsx."""

from __future__ import annotations

import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from core.utils import format_date_dd_mm_yyyy, normalize_text

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TEMPLATE_PATH = ROOT / "template.xlsx"
SETUP_KEY_TEMPLATE_PATH = "supplier_template_path"

# Hàng / cột theo template (1-based)
ROW_PROPOSER = 5
ROW_SUPPLIER = 6
ROW_REASON = 7
ROW_HEADER = 8
DATA_START_ROW = 9
COL_STT = 1
COL_MATERIAL = 2
COL_PRODUCT = 3
COL_DG = 4
COL_COLOR = 6
COL_LOGO = 8
COL_QTY = 10
COL_DETAIL = 11
COL_CREATED_DATE = 11  # K5

DEFAULT_REASON = "giao lần đầu"
TEMPLATE_DATA_ROWS = 6  # số dòng mẫu trong template (9-14)
TEMPLATE_TOTAL_ROW = 15
TEMPLATE_SIG_ROW = 19


def resolve_supplier_template_path(
    db=None,
    *,
    explicit_path: str | Path | None = None,
) -> Path:
    """Setup path → explicit → template.xlsx mặc định trong project."""
    if explicit_path:
        p = Path(explicit_path)
        if p.is_file():
            return p.resolve()
        raise FileNotFoundError(f"Không tìm thấy template: {p}")

    if db is not None:
        saved = normalize_text(db.get_setup(SETUP_KEY_TEMPLATE_PATH, ""))
        if saved:
            p = Path(saved)
            if p.is_file():
                return p.resolve()
            raise FileNotFoundError(
                f"Template trong Setup không tồn tại: {p}\n"
                f"Vào Setup → Tài khoản → chọn lại file mẫu phiếu Supplier."
            )

    if DEFAULT_TEMPLATE_PATH.is_file():
        return DEFAULT_TEMPLATE_PATH.resolve()
    raise FileNotFoundError(
        f"Không tìm thấy template mặc định: {DEFAULT_TEMPLATE_PATH}\n"
        "Vào Setup → Tài khoản → Template phiếu Supplier để chọn file .xlsx."
    )


def _copy_row_style(ws, src_row: int, dest_row: int, max_col: int = 11) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dest_row, column=col)
        dst._style = copy.copy(src._style)
        dst.font = copy.copy(src.font)
        dst.border = copy.copy(src.border)
        dst.fill = copy.copy(src.fill)
        dst.number_format = copy.copy(src.number_format)
        dst.protection = copy.copy(src.protection)
        dst.alignment = copy.copy(src.alignment)


def export_slip_to_excel(
    slip: dict[str, Any],
    dest_path: str,
    *,
    db=None,
    template_path: str | Path | None = None,
) -> str:
    tpl = resolve_supplier_template_path(db, explicit_path=template_path)

    lines = list(slip.get("lines") or [])
    line_count = len(lines)
    extra = max(0, line_count - TEMPLATE_DATA_ROWS)

    wb = load_workbook(tpl)
    ws = wb.active

    if extra > 0:
        ws.insert_rows(TEMPLATE_TOTAL_ROW, amount=extra)
        for i in range(extra):
            dest = DATA_START_ROW + TEMPLATE_DATA_ROWS + i
            _copy_row_style(ws, DATA_START_ROW, dest)

    total_row = TEMPLATE_TOTAL_ROW + extra
    sig_row = TEMPLATE_SIG_ROW + extra

    ws.cell(row=ROW_PROPOSER, column=5, value=normalize_text(slip.get("proposed_by")))
    ws.cell(row=ROW_SUPPLIER, column=5, value=normalize_text(slip.get("supplier")))
    reason = normalize_text(slip.get("reason")) or DEFAULT_REASON
    ws.cell(row=ROW_REASON, column=5, value=reason)

    created = slip.get("created_at", "")
    try:
        from datetime import datetime

        dt = datetime.fromisoformat(str(created).replace("Z", ""))
        created_txt = format_date_dd_mm_yyyy(dt)
    except (ValueError, TypeError):
        created_txt = str(created)[:10]
    ws.cell(row=ROW_PROPOSER, column=COL_CREATED_DATE, value=created_txt)

    qty_total = 0.0
    for i, line in enumerate(lines):
        r = DATA_START_ROW + i
        ws.cell(row=r, column=COL_STT, value=int(line.get("line_no") or i + 1))
        ws.cell(row=r, column=COL_MATERIAL, value=normalize_text(line.get("material_code")))
        ws.cell(row=r, column=COL_PRODUCT, value=normalize_text(line.get("product_code")))
        ws.cell(row=r, column=COL_DG, value=normalize_text(line.get("dg_case")))
        ws.cell(row=r, column=COL_COLOR, value=normalize_text(line.get("color")))
        ws.cell(row=r, column=COL_LOGO, value=normalize_text(line.get("logo")))
        try:
            q = float(line.get("quantity") or 0)
        except (TypeError, ValueError):
            q = 0.0
        qty_total += q
        if q == int(q):
            ws.cell(row=r, column=COL_QTY, value=int(q))
        else:
            ws.cell(row=r, column=COL_QTY, value=q)
        ws.cell(row=r, column=COL_DETAIL, value=normalize_text(line.get("detail")))

    for r in range(DATA_START_ROW + line_count, total_row):
        for c in range(1, 12):
            ws.cell(row=r, column=c, value=None)

    ws.cell(row=total_row, column=COL_QTY, value=int(qty_total) if qty_total == int(qty_total) else qty_total)
    if total_row + 1 <= sig_row - 1:
        ws.cell(row=total_row + 1, column=COL_QTY, value=int(qty_total) if qty_total == int(qty_total) else qty_total)

    out = Path(dest_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return str(out.resolve())
