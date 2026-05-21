"""Bảng kê định mức sản xuất — đọc/cache theo hash ô A6."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pandas as pd

from core.database import HubDatabase
from core.utils import (
    compute_file_signature,
    extract_customer_code_from_product_code,
    extract_item_code_from_product_code,
    hash_text,
    looks_like_dg_case,
    normalize_text,
)

BOM_KE_HEADER_ROW = 8
BOM_KE_DATA_START_ROW = 9
BOM_KE_A6_CELL = "A6"


@dataclass
class BomKeLoadResult:
    source: str  # "cache_a6" | "parsed"
    file_path: str
    file_name: str
    file_hash: str
    a6_text: str
    a6_hash: str
    row_count: int
    message: str
    df: pd.DataFrame


class BomKeReaderService:
    def __init__(self, db: HubDatabase | None = None):
        self.db = db or HubDatabase()

    def read_a6_text(self, file_path: str) -> str:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Không tìm thấy bảng kê: {file_path}")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            return normalize_text(ws[BOM_KE_A6_CELL].value)
        finally:
            wb.close()

    def parse_bom_ke_excel(self, file_path: str) -> tuple[str, str, pd.DataFrame]:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Không tìm thấy bảng kê: {file_path}")

        a6_text = self.read_a6_text(file_path)
        if not a6_text:
            raise ValueError("Ô A6 trống — không xác định được phiên bản bảng kê.")
        a6_hash = hash_text(a6_text)

        df_raw = pd.read_excel(file_path, sheet_name=0, header=None)
        if df_raw.empty:
            raise ValueError("Bảng kê trống.")
        if df_raw.shape[1] < 16:
            raise ValueError("Bảng kê không đủ cột (cần tối thiểu cột P).")

        data = df_raw.iloc[BOM_KE_DATA_START_ROW - 1 :].copy()
        out = pd.DataFrame(
            {
                "row_index": data.index + 1,
                "dg_case": data.iloc[:, 0].map(normalize_text),
                "order_date": pd.to_datetime(data.iloc[:, 1], errors="coerce"),
                "product_code": data.iloc[:, 3].map(normalize_text),
                "qty_divisor": pd.to_numeric(data.iloc[:, 6], errors="coerce"),
                "ma_npl": data.iloc[:, 9].map(normalize_text),
                "ten_npl": data.iloc[:, 10].map(normalize_text),
                "mo_ta": data.iloc[:, 11].map(normalize_text),
                "don_vi_tinh": data.iloc[:, 13].map(normalize_text),
                "so_luong_dm_1": pd.to_numeric(data.iloc[:, 14], errors="coerce"),
                "so_luong": pd.to_numeric(data.iloc[:, 15], errors="coerce"),
            }
        )
        out["customer_code"] = out["product_code"].map(extract_customer_code_from_product_code)
        out["item_code"] = out["product_code"].map(extract_item_code_from_product_code)
        out = out[
            (out["dg_case"] != "")
            & (out["product_code"] != "")
            & out["dg_case"].map(looks_like_dg_case)
        ].copy()
        out.reset_index(drop=True, inplace=True)
        return a6_text, a6_hash, out

    def load(self, file_path: str, *, force: bool = False) -> BomKeLoadResult:
        path = str(Path(file_path).resolve())
        file_name = Path(path).name
        file_hash = compute_file_signature(path)
        a6_text = self.read_a6_text(path)
        a6_hash = hash_text(a6_text)

        if not force:
            meta = self.db.get_bom_ke_dataset(a6_hash)
            if meta and str(meta["file_hash"]) == file_hash:
                df = self.db.load_bom_ke_dataset_df(a6_hash)
                if df is not None:
                    return BomKeLoadResult(
                        source="cache_a6",
                        file_path=path,
                        file_name=file_name,
                        file_hash=file_hash,
                        a6_text=a6_text,
                        a6_hash=a6_hash,
                        row_count=len(df),
                        message=f"Dùng cache bảng kê theo A6 ({len(df)} dòng).",
                        df=df,
                    )

        a6_text, a6_hash, df = self.parse_bom_ke_excel(path)
        self.db.save_bom_ke_dataset(path, file_hash, a6_text, a6_hash, df)
        self.db.set_setup("bom_ke_file_path", path)
        self.db.set_setup("bom_ke_a6_hash", a6_hash)
        self.db.set_setup("bom_ke_a6_text", a6_text)

        return BomKeLoadResult(
            source="parsed",
            file_path=path,
            file_name=file_name,
            file_hash=file_hash,
            a6_text=a6_text,
            a6_hash=a6_hash,
            row_count=len(df),
            message=f"Đã đọc {len(df)} dòng bảng kê — cache theo A6 ({a6_hash[:8]}…).",
            df=df,
        )

    def load_cached(self, a6_hash: str | None = None) -> BomKeLoadResult | None:
        key = a6_hash or self.db.get_setup("bom_ke_a6_hash", "")
        if not key:
            return None
        meta = self.db.get_bom_ke_dataset(key)
        if not meta:
            return None
        df = self.db.load_bom_ke_dataset_df(key)
        if df is None:
            return None
        return BomKeLoadResult(
            source="cache_a6",
            file_path=str(meta["file_path"]),
            file_name=str(meta["file_name"]),
            file_hash=str(meta["file_hash"]),
            a6_text=str(meta["a6_text"]),
            a6_hash=str(meta["a6_hash"]),
            row_count=len(df),
            message=f"Đã tải cache bảng kê ({len(df)} dòng).",
            df=df,
        )

    def query_by_dg_case(self, dg_case: str, *, a6_hash: str | None = None) -> pd.DataFrame:
        key = a6_hash or self.db.get_setup("bom_ke_a6_hash", "")
        if not key:
            return pd.DataFrame()
        return self.db.query_bom_ke_rows(key, dg_case=dg_case)

    def subset_by_dg_case(self, df: pd.DataFrame, dg_case: str) -> pd.DataFrame:
        from core.utils import normalize_dg_case

        key = normalize_dg_case(dg_case)
        if not key or df.empty:
            return df.iloc[0:0].copy()
        mask = df["dg_case"].map(
            lambda x: key == normalize_dg_case(x) or key in normalize_dg_case(x)
        )
        subset = df[mask].copy()
        if subset.empty:
            return subset
        base_product_code = normalize_text(subset.iloc[0]["product_code"])
        from core.utils import normalize_key

        return df[df["product_code"].map(normalize_key) == normalize_key(base_product_code)].copy()
